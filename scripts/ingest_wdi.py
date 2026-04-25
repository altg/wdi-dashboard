"""
One-time ingestion: WDIEXCEL.xlsx -> wdidata/wdi.duckdb

Run from the wdi-dashboard directory:
    python scripts/ingest_wdi.py

Requires: pip install openpyxl duckdb pandas
Output:   wdidata/wdi.duckdb  (opened read-only by the Next.js app)
"""

import sys
import os
import time
import duckdb
import openpyxl
import pandas as pd

EXCEL_PATH = "wdidata/WDIEXCEL.xlsx"
DB_PATH = "wdidata/wdi.duckdb"

REGION_CODES = {
    "East Asia & Pacific":        "EAS",
    "Europe & Central Asia":      "ECA",
    "Latin America & Caribbean":  "LAC",
    "Middle East & North Africa": "MNA",
    "North America":              "NAC",
    "South Asia":                 "SAS",
    "Sub-Saharan Africa":         "SSA",
}

INCOME_CODES = {
    "High income":         "HIC",
    "Low income":          "LIC",
    "Lower middle income": "LMC",
    "Upper middle income": "UMC",
}


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: {EXCEL_PATH} not found. Run from the wdi-dashboard directory.")
        sys.exit(1)

    for f in [DB_PATH, DB_PATH + ".wal"]:
        if os.path.exists(f):
            os.remove(f)
            print(f"Removed {f}")

    print(f"Reading {EXCEL_PATH} ...")
    t0 = time.time()
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"  Workbook opened in {time.time() - t0:.1f}s")

    # ── Country sheet ──────────────────────────────────────────────────────────
    print("Parsing Country sheet ...")
    country_rows = []
    for i, row in enumerate(wb["Country"].iter_rows(values_only=True)):
        if i == 0:
            continue
        code, name = row[0], row[1]
        region_name, income_name = row[7], row[8]
        if not code:
            continue
        country_rows.append({
            "country_code":  str(code),
            "country_name":  str(name) if name else "",
            "region_code":   REGION_CODES.get(region_name, "") if region_name else "",
            "region_name":   str(region_name) if region_name else "",
            "income_code":   INCOME_CODES.get(income_name, "") if income_name else "",
            "income_name":   str(income_name) if income_name else "",
        })
    countries_df = pd.DataFrame(country_rows)
    print(f"  {len(countries_df)} countries parsed")

    # ── Data sheet (wide -> long) ───────────────────────────────────────────────
    print("Parsing Data sheet (wide to long) ...")
    ws_data = wb["Data"]
    year_cols: list[tuple[int, int]] = []
    obs_rows = []
    source_count = 0

    t1 = time.time()
    for i, row in enumerate(ws_data.iter_rows(values_only=True)):
        if i == 0:
            for j, h in enumerate(row):
                if h and str(h).isdigit():
                    year_cols.append((j, int(h)))
            print(f"  Year range: {year_cols[0][1]}-{year_cols[-1][1]} ({len(year_cols)} years)")
            continue

        source_count += 1
        country_code   = row[1]
        country_name   = row[0]
        indicator_code = row[3]
        if not country_code or not indicator_code:
            continue

        cc = str(country_code)
        cn = str(country_name) if country_name else ""
        ic = str(indicator_code)

        for j, year in year_cols:
            val = row[j]
            if val is None:
                continue
            try:
                obs_rows.append((cc, cn, ic, year, float(val)))
            except (TypeError, ValueError):
                pass

        if i % 20_000 == 0 and i > 0:
            print(f"  {source_count:>6} rows ... {len(obs_rows):,} obs ({time.time()-t1:.0f}s)")

    t_parse = time.time() - t1
    print(f"  {source_count:,} source rows -> {len(obs_rows):,} non-null obs ({t_parse:.1f}s)")

    # ── Build DataFrame and write to DuckDB ────────────────────────────────────
    print("Building DataFrame ...")
    t2 = time.time()
    obs_df = pd.DataFrame(obs_rows, columns=["country_code", "country_name", "indicator_code", "year", "value"])
    obs_df["year"] = obs_df["year"].astype("int16")
    obs_df["value"] = obs_df["value"].astype("float64")
    print(f"  DataFrame: {len(obs_df):,} rows, {obs_df.memory_usage(deep=True).sum() / 1e6:.0f} MB ({time.time()-t2:.1f}s)")

    print(f"Writing {DB_PATH} ...")
    t3 = time.time()
    con = duckdb.connect(DB_PATH)

    # Countries table
    con.execute("""
        CREATE TABLE countries (
            country_code  VARCHAR NOT NULL PRIMARY KEY,
            country_name  VARCHAR NOT NULL,
            region_code   VARCHAR NOT NULL,
            region_name   VARCHAR NOT NULL,
            income_code   VARCHAR NOT NULL,
            income_name   VARCHAR NOT NULL
        )
    """)
    con.execute("INSERT INTO countries SELECT * FROM countries_df")
    print(f"  countries: {len(countries_df)} rows ({time.time()-t3:.1f}s)")

    # Observations table — DuckDB queries the DataFrame directly (fast)
    t4 = time.time()
    con.execute("""
        CREATE TABLE observations AS
        SELECT country_code, country_name, indicator_code,
               CAST(year AS SMALLINT) AS year,
               value
        FROM obs_df
    """)
    print(f"  observations: {len(obs_df):,} rows ({time.time()-t4:.1f}s)")

    print("  Building indexes ...")
    t5 = time.time()
    con.execute("CREATE INDEX idx_ind_country ON observations(indicator_code, country_code)")
    con.execute("CREATE INDEX idx_ind_year    ON observations(indicator_code, year)")
    con.execute("CHECKPOINT")
    con.close()
    print(f"  Indexes + checkpoint ({time.time()-t5:.1f}s)")

    db_mb = os.path.getsize(DB_PATH) / 1024 / 1024
    print(f"\nDone! {DB_PATH} ({db_mb:.0f} MB) in {time.time()-t0:.0f}s total")
    print("Set DATA_SOURCE=local in .env.local to activate local mode.")


if __name__ == "__main__":
    main()
